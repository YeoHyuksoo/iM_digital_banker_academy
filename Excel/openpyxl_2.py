from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("diary", 0)

data = [('홍길동', 80, 70, 90), ('이기자', 90, 60, 80), ('강기자', 80, 80, 70)]

i = 1
start_col = 'D'
for d in data:
    for j in range(4):
        #ws.cell(row = i, column = j+1).value = d[j]
        ws[chr(ord(start_col) + j) + str(i)].value = d[j]
    i += 1

ws.cell(row = i, column = 1).value = '합계'
ws.cell(row = i, column = 2).value = '=sum(B1:B3)'
ws.cell(row = i, column = 3).value = '=sum(C1:C3)'
ws.cell(row = i, column = 4).value = '=sum(D1:D3)'

wb.save("openpyxl_2.xlsx")
wb.close()

