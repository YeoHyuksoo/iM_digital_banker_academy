from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "sample"
ws['B2'] = 42
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
wb.save("openpyxl_1.xlsx")

wb.close()

wb = load_workbook(filename="openpyxl_1.xlsx")
#ws = wb.active
ws = wb["sample"]

ws['A1'] = 42
ws.cell(row = 1, column = 3).value = 333
ws.append(['aaa', 'bbb', 'ccc'])

print(ws['A1'].value)

ws2 = wb['sample']
print(ws2['C3'].value)
print(ws.cell(row = 1, column = 3).value)

wb.save("openpyxl_1.xlsx")
wb.close()